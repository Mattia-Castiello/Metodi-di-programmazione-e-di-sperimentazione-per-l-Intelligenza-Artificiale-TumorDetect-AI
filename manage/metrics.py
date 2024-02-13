import os
import numpy as np
from openpyxl import load_workbook
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt

class Metrics:
    """
    Modella le metriche di validazione del modello
    """
    def __init__(self, true_positive, true_negative, false_positive, false_negative, metrics=[], filename=None):
        """
        Costruttore

        Parameters
        ----------
        true_positive : int
            il numero di veri positivi
        true_negative : int
            il numero di veri negativi
        false_positive : int
            il numero di falsi positivi
        false_negative : int
            il numero di falsi negativi
        metrics : list
            la lista delle metriche

        Returns
        -------
        None
        """
        self.true_positive = true_positive
        self.true_negative = true_negative
        self.false_positive = false_positive
        self.false_negative = false_negative
        self.metrics = metrics
        self.filename = filename


    def confusion_matrix(self):
        """
        Calcola la confusion matrix

        Returns
        -------
        list
            la confusion matrix
        """
        confusion_matrix = [self.true_negative, self.false_positive, self.false_negative, self.true_positive]

        return confusion_matrix


    def accuracy(self, confusion_matrix, K=1):
        """
        Calcola l'accuratezza

        Parameters
        ----------
        confusion_matrix : list
            la confusion matrix

        K : int
            il numero di esperimenti
        Returns
        -------
        accuracy : float
            l'accuratezza

        accuracy scores: list
            i valori di accuratezza per ogni esperimento

        """
        accuracy_scores = []
        true_positive = confusion_matrix[0]
        false_positive = confusion_matrix[1]
        false_negative = confusion_matrix[2]
        true_negative = confusion_matrix[3]

        # Se la confusion matrix è una lista di liste
        if isinstance(confusion_matrix[0], list):
            for i in range(K):
                accuracy_scores.append((true_positive[i] + true_negative[i]) / (true_positive[i] + false_positive[i] + false_negative[i] + true_negative[i]))
        # Se la confusion matrix è una lista singola
        else:
            accuracy_scores.append((true_positive + true_negative) / (true_positive + false_positive + false_negative + true_negative))

        # Calcola l'accuratezza media
        accuracy = np.mean(accuracy_scores)
        return accuracy, accuracy_scores


    def error_rate(self, confusion_matrix, K=1):
        """
        Calcola l'error rate

        Parameters
        ----------
        confusion_matrix : list
            la confusion matrix

        K : int
            il numero di esperimenti
        Returns
        -------
        error_rate : float
            l'error rate

        error_rate_scores : list
            i valori di error rate per ogni esperimento

        """
        error_rate_scores = []
        true_positive = confusion_matrix[0]
        false_positive = confusion_matrix[1]
        false_negative = confusion_matrix[2]
        true_negative = confusion_matrix[3]

        # Se la confusion matrix è una lista di liste
        if isinstance(confusion_matrix[0], list):
            for i in range(K):
                error_rate_scores.append((false_positive[i] + false_negative[i]) / (true_positive[i] + false_positive[i] + false_negative[i] + true_negative[i]))
        # Se la confusion matrix è una lista singola
        else:
            error_rate_scores.append((false_positive + false_negative) / (true_positive + false_positive + false_negative + true_negative))

        # Calcola l'error rate medio
        error_rate = np.mean(error_rate_scores)
        return error_rate, error_rate_scores


    def sensitivity(self, confusion_matrix, K=1):
        """
        Calcola la sensitivity

        Parameters
        ----------
        confusion_matrix : list
            la confusion matrix

        K : int
            il numero di esperimenti
        Returns
        -------
        sensitivity : float
            la sensitivity

        sensitivity_scores : list
            i valori di sensitivity per ogni esperimento

        """
        sensitivity_scores = []
        true_positive = confusion_matrix[0]
        false_positive = confusion_matrix[1]
        false_negative = confusion_matrix[2]
        true_negative = confusion_matrix[3]

        # Se la confusion matrix è una lista di liste
        if isinstance(confusion_matrix[0], list):
            for i in range(K):
                sensitivity_scores.append(true_positive[i] / (true_positive[i] + false_negative[i]))
        # Se la confusion matrix è una lista singola
        else:
            sensitivity_scores.append(true_positive / (true_positive + false_negative))

        sensitivity = np.mean(sensitivity_scores)
        return sensitivity, sensitivity_scores


    def specificity(self, confusion_matrix, K=1):
        """
        Calcola la specificity

        Parameters
        ----------
        confusion_matrix : list
            la confusion matrix

        K : int
            il numero di esperimenti
        Returns
        -------
        specificity : float
            la specificity

        specificity_scores : list
            i valori di specificity per ogni esperimento

        """
        specificity_scores = []
        true_positive = confusion_matrix[0]
        false_positive = confusion_matrix[1]
        false_negative = confusion_matrix[2]
        true_negative = confusion_matrix[3]

        # Se la confusion matrix è una lista di liste
        if isinstance(confusion_matrix[0], list):
            for i in range(K):
                specificity_scores.append(true_negative[i] / (true_negative[i] + false_positive[i]))
        # Se la confusion matrix è una lista singola
        else:
            specificity_scores.append(true_negative / (true_negative + false_positive))

        # Calcola la specificity media
        specificity = np.mean(specificity_scores)
        return specificity, specificity_scores


    def geometric_mean(self, confusion_matrix, K=1):
        """
        Calcola la geometric mean

        Parameters
        ----------
        confusion_matrix : list
            la confusion matrix

        K : int
            il numero di esperimenti
        Returns
        -------
        geometric_mean : float
            la geometric mean

        geometric_mean_scores : list
            i valori di geometric mean per ogni esperimento

        """
        geometric_mean_scores = []
        true_positive = confusion_matrix[0]
        false_positive = confusion_matrix[1]
        false_negative = confusion_matrix[2]
        true_negative = confusion_matrix[3]

        # Se la confusion matrix è una lista di liste
        if isinstance(confusion_matrix[0], list):
            for i in range(K):
                sensitivity = true_positive[i] / (true_positive[i] + false_negative[i])
                specificity = true_negative[i] / (true_negative[i] + false_positive[i])
                geometric_mean_scores.append(np.sqrt(sensitivity * specificity))
        # Se la confusion matrix è una lista singola
        else:
            sensitivity = true_positive / (true_positive + false_negative)
            specificity = true_negative / (true_negative + false_positive)
            geometric_mean_scores.append(np.sqrt(sensitivity * specificity))

        # Calcola la geometric mean media
        geometric_mean = np.mean(geometric_mean_scores)
        return geometric_mean, geometric_mean_scores


    def get_metrics(self, K=1):
        """
        Calcola tutte le metriche

        Parameters
        ----------

        K : int
            il numero di esperimenti
        Returns
        -------
        list
            le metriche
        """

        metrics = {}
        confusion_matrix = self.confusion_matrix()

        # Calcola le metriche richieste
        if 1 in self.metrics:
            metrics['Accuracy'] = self.accuracy(confusion_matrix, K)
        if 2 in self.metrics:
            metrics['Error Rate'] = self.error_rate(confusion_matrix, K)
        if 3 in self.metrics:
            metrics['Sensitivity'] = self.sensitivity(confusion_matrix, K)
        if 4 in self.metrics:
            metrics['Specificity'] = self.specificity(confusion_matrix, K)
        if 5 in self.metrics:
            metrics['Geometric Mean'] = self.geometric_mean(confusion_matrix, K)
        if 6 in self.metrics:
            metrics = {
                'Accuracy': self.accuracy(confusion_matrix, K),
                'Error Rate': self.error_rate(confusion_matrix, K),
                'Sensitivity': self.sensitivity(confusion_matrix, K),
                'Specificity': self.specificity(confusion_matrix, K),
                'Geometric Mean': self.geometric_mean(confusion_matrix, K)
            }
        return metrics


    def save_metrics(self, metrics, filename=None):
        """
        Salva le metriche su file Excel

        Parameters
        ----------
        metrics : dict
            le metriche

        filename : str
            il nome del file Excel
        Returns
        -------
        None
        """
        if filename is None:
            filename = input("Inserisci il nome del file Excel (senza estensione .xlsx): ")
            filename = filename + '.xlsx' if '.xlsx' not in filename else filename # Aggiungi l'estensione .xlsx se non è presente
        
        self.filename = filename

        metric_values = []
        max_length = 0 # Lunghezza massima della colonna dei valori
        for metric, value in metrics.items():
            # Verifica se il valore è una lista
            if isinstance(value, list):
                # Se è una lista, formatta la lista come stringa
                value_str = ', '.join([str(v) for v in value])
            else:
                value_str = value[1]
            metric_values.append((metric, value_str))
            max_length = max(max_length, len(str(value_str))) # Aggiorna la lunghezza massima 

        df = pd.DataFrame(metric_values, columns=['Metriche', 'Valori'])
        df.to_excel(filename, sheet_name='Risultati Metriche', index=False)

        # Imposta la larghezza delle colonne utilizzando openpyxl
        workbook = load_workbook(filename)
        sheet = workbook['Risultati Metriche']
        sheet.column_dimensions['A'].width = 20  # Imposta la larghezza della colonna A
        sheet.column_dimensions['B'].width = max_length + 10  # Imposta la larghezza della colonna B
        workbook.save(filename)

        print("Le metriche sono state salvate su file Excel.")


    def save_plot_to_excel(self, temp_file, sheet_name):
        """
        Salva un plot in un foglio Excel esistente o nuovo.

        Parameters
        ----------
        temp_file : str
            Il nome dell'immagine temporanea del plot.

        sheet_name : str
            Il nome del foglio in cui salvare il plot.

        Returns
        -------
        None
        """
        try:
            # Carica il foglio Excel
            wb = load_workbook(self.filename)
            try:
                ws = wb[sheet_name]
            except KeyError:
                # Se il foglio non esiste, crea un nuovo foglio
                ws = wb.create_sheet(title=sheet_name)

            # Aggiungi il plot al foglio Excel
            img = openpyxl.drawing.image.Image(temp_file)
            img.width = img.width * 0.8  # Riduci la larghezza dell'immagine del 20%
            img.height = img.height * 0.8  # Riduci l'altezza dell'immagine del 20%
            ws.add_image(img, 'A1')

            # Salva le modifiche
            wb.save(self.filename)

            # Elimina il file temporaneo
            try:
                os.remove(temp_file)
            except OSError as e:
                print(f"Errore durante l'eliminazione del file temporaneo del plot.")
        except Exception as e:
            print(f"Errore durante il salvataggio del plot su Excel.")


    def metrics_plot(self, metrics):
        """
        Mostra i grafici delle metriche

        Parameters
        ----------
        metrics : dict
            le metriche
        Returns
        -------
        None
        """
        # Estraiamo le etichette delle metriche
        labels = list(metrics.keys())
        # Estraiamo le liste dei valori delle metriche
        metric_scores = [value[1] for value in metrics.values()]

        # Plot del boxplot
        plt.figure(figsize=(12, 6))
        plt.boxplot(metric_scores, labels=labels)

        # Asse x
        plt.xlabel('Metriche')
        # Asse y
        plt.ylabel('Valori')
        # Titolo del grafico
        plt.title('Metriche di validazione')
        # Mostra la griglia
        plt.grid(True)

        # Salva il plot in un file temporaneo
        temp_file = 'temp_plot1.png'
        plt.savefig(temp_file)
        self.save_plot_to_excel(temp_file, 'Plot Andatamento Metriche')

        # Estrazione dei dati per il plot a linea
        K_experiments = list(range(1, len(metric_scores[0]) + 1))

        # Plot per ciascuna metrica
        plt.figure(figsize=(12, 6))
        # Plot per ciascuna metrica
        for label, values in zip(labels, metric_scores):
            plt.plot(K_experiments, values, marker='o', label=label)

        # Etichette e titoli
        plt.xlabel('Esperimento K')
        plt.ylabel('Valore della metrica')
        plt.title('Andamento delle metriche negli esperimenti K')

        # Mostra la legenda
        plt.legend()

        # Mostra il grafico
        plt.grid(True)

        # Salva il plot in un file temporaneo
        temp_file = 'temp_plot2.png'
        plt.savefig(temp_file)
        self.save_plot_to_excel(temp_file, 'Andamento Metriche in K')

        # Mostra il plot
        plt.show()




